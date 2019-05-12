import openpyxl as px
import json

specialities_dict = {
    "behind_1":"ビハインド◯", "behind_2":"ビハインド◎", "lead_1":"リード◯",
    "lead_2":"リード◎", "physical_1":"フィジカル◯", "physical_2":"フィジカル◎",
    "anti_inj_1":"ケガしにくさ◯", "anti_inj_2":"ケガしにくさ◎", "dash":"ダッシュ◯",
    "dribbler_1":"ドリブラー◯", "dribbler_2":"ドリブラー◎", "onetouch":"ワンタッチ◯",
    "trap_1":"トラップ◯", "trap_2":"トラップ◎", "air_1":"空中戦◯", "air_2":"空中戦◎",
    "bt_1":"突破力◯", "bt_2":"突破力◎", "kp_1":"キープ力◯", "kp_2":"キープ力◎",
    "pass_1":"パス◯", "pass_2":"パス◎", "tpass":"スルーパス◯", "cross":"クロス◯",
    "expansion":"展開力◯", "backspin":"バックスピン◯", "assist":"アシスト◯",
    "pass_intuition":"パス勘◯", "shoot_1":"シュート力◯", "shoot_2":"シュート力◎",
    "s_range":"シュートレンジ◯", "d_force_1":"決定力◯", "d_force_2":"決定力◎",
    "heading":"ヘディング◯", "mark":"マーク外し", "matchup":"マッチアップ◯",
    "st_ball_1":"ボール奪取◯", "st_ball_2":"ボール奪取◎", "covering":"カバーリング◯",
    "danger_sense":"危機察知◯", "def_reaction":"守備反応◯", "recapture":"奪還◯",
    "prefetch":"先読み", "coaching":"コーチング", "free_kick":"フリーキック◯",
    "pk":"PK◯", "long_throw":"ロングスロー", "teem_play":"チームプレー◯",
    "mood":"ムード◯", "f_split":"闘争心", "explo_pow":"爆発力◯", "guts_1":"根性◯",
    "guts_2":"根性◎", "recover_1":"回復◯", "recover_2":"回復◎", "hardwork":"ハードワーク◯",
    "playmaker":"司令塔", "captaincy":"キャプテンシー", "counter":"カウンター◯",
    "eos":"意外性", "vs_ace":"対エース◯", "fair_play":"フェアプレー",
    "malicia":"マリーシア", "big_match_1":"大一番◯", "big_match_2":"大一番◎",
    "first_goal":"ファーストゴール", "bench":"途中出場◯", "final_phase":"終盤◯",
    "lucky_boy":"ラッキーボーイ", "soccer_it_1":"サッカー脳◯", "soccer_it_2":"サッカー脳◎",
    "saving_1":"セービング◯", "saving_2":"セービング◎", "highball":"ハイボールキャッチ",
    "cocentration":"集中力", "low_punt_kick":"低弾道パントキック", "cor_avoid_1":"コーナー回避◯",
    "cor_avoid_2":"コーナー回避◎", "zero":"無失点", "vs_long_shoot":"対ロングシュート◯"
}

wb = px.load_workbook("speciality.xlsx")
ws = wb.active
start = 2
end = 115
specialities = []
for i in range(start,end):
    num = str(i)
    lst = {
        "name": str(ws["A"+num].value),
        "tag": "tag",
        "physical": str(ws["B"+num].value),
        "speed": str(ws["C"+num].value),
        "tecnic": str(ws["D"+num].value),
        "mental": str(ws["E"+num].value),
        "cf": str(ws["F"+num].value),
        "wg": str(ws["G"+num].value),
        "st": str(ws["H"+num].value),
        "omf": str(ws["I"+num].value),
        "cmf": str(ws["J"+num].value),
        "smf": str(ws["K"+num].value),
        "dmf": str(ws["L"+num].value),
        "cb": str(ws["M"+num].value),
        "sb": str(ws["N"+num].value),
        "gk": str(ws["O"+num].value),
    }
    for i, j in specialities_dict.items():
        if lst["name"] == j:
            lst["tag"] = i
            break
        else:
            continue
    #specialities[num] = lst
    specialities.append(lst)
#with open("specialities.json", mode="w") as f:
#    json.dump(specialities,f)

with open("specialities.txt", mode="w") as f:
    f.write("[")
    for i in specialities:
        f.write("{")
        for j, k in i.items():
            f.write('"%s":"%s",' % (j,k))
        else:
            f.write("},")
            f.write("\n")
    else:
        f.write("]")

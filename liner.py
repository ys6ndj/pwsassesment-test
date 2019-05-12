import pulp
import openpyxl as px



WHATS_POSITION = "OMF"
wb = px.load_workbook("speciality.xlsx")
ws = wb.active
start = 2
end = 115
specialities = []
for i in range(start, end):
    num = str(i)
    lst = [
        str(ws["A"+num].value),
        float(str(ws["B"+num].value)),
        float(str(ws["C"+num].value)),
        float(str(ws["D"+num].value)),
        float(str(ws["E"+num].value)),
        float(str(ws["I"+num].value)),
    ]
    specialities.append(lst)
wb.save("speciality.xlsx")
for i in specialities:
    print(i)
num_specialities = len(specialities)

prob = pulp.LpProblem('selection problem', pulp.LpMaximize)
print(prob)
x = [
    pulp.LpVariable('x[{}]'.format(i), 0, 1, 'Integer') for i in range(num_specialities)
]

obj = 0
for i in range(num_specialities):
    obj += specialities[i][5] * x[i]

prob += obj
print(prob)


fi, sp, te, he = 0, 0, 0, 0
for i in range(num_specialities):
    fi += specialities[i][1] * x[i]
    sp += specialities[i][2] * x[i]
    te += specialities[i][3] * x[i]
    he += specialities[i][4] * x[i]

prob += (fi <= 999)
prob += (sp <= 999)
prob += (te <= 999)
prob += (he <= 999)

print(prob)

status = prob.solve()
print('Status', pulp.LpStatus[status])


total = 0
lst = []
er_lst = []
for i, speciality in enumerate(specialities):
    print((pulp.value(x[i])))
    print(type(pulp.value(x[i])))
    try:
        cnt = int(pulp.value(x[i]))
    except:
        cnt = 0.0
        er_lst.append(speciality)
    print('{} ({}, {}, {}, {}, {}): {}'.format(*speciality, cnt))
    total += speciality[5] * cnt
    if cnt:
        lst.append([*speciality, cnt])
print('total', total)
print(lst)
t = 0
for i in lst:
    t += i[-2]
print(t)
print(er_lst)
